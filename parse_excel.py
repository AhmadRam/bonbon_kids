import json
try:
    import pandas as pd
    df = pd.read_excel('D:\\laragon\\www\\bagisto-2.4\\عروض.xlsx')
    with open('excel_data.json', 'w', encoding='utf-8') as f:
        f.write(df.to_json(orient='records', force_ascii=False))
except ImportError:
    try:
        from openpyxl import load_workbook
        wb = load_workbook('D:\\laragon\\www\\bagisto-2.4\\عروض.xlsx', data_only=True)
        sheet = wb.active
        data = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
            data.append(row_data)
        with open('excel_data.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        with open('excel_data.json', 'w', encoding='utf-8') as f:
            f.write(str(e))
